import os
import re
import pandas as pd
import myFunctions
import math
import json
import numpy as np
from scipy.optimize import curve_fit
import statsmodels.api as sm
import statsmodels.formula.api as smf
import datetime
from openpyxl import Workbook, load_workbook
from datetime import date, timedelta

# windows 获取当前文件所在目录
# current_file_path = os.path.abspath(__file__)
# pDir = os.path.dirname(current_file_path)

#linux-获取当前目录
pDir = os.getcwd()

main_excel_file = os.path.join(pDir, 'HIC/ALL_Tweets_list_HIC_KeyInfo.xlsx')
hic_minute_start = '2023-10-06 08:00:00'
hic_minute_end = '2023-10-31 07:59:59'

#accountID的首字母集合为：{0~9, A~Z, _, a~z}，共63个
accountID_init = [
    ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', 'A', 'B', 'C'],
    ['D', 'E', 'F', 'G', 'H'],
    ['I', 'J', 'K', 'L', 'M'],
    ['N', 'O', 'P', 'Q', 'R', 'S'],
    ['T', 'U', 'V', 'W', 'X', 'Y', 'Z', '_', 'a'],
    ['b', 'c', 'd', 'e', 'f', 'g', 'h', 'i'],
    ['j', 'k', 'l', 'm', 'n', 'o', 'p'],
    ['q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']
]

def get_minute_hour_list(minute_start, minute_end):
    # 生成分钟列表
    minute_list = pd.date_range(start=minute_start, end=minute_end, freq='min').strftime('%Y-%m-%d %H:%M').tolist()
    # 将结束时间加一小时，保证包含最后一个小时
    new_minute_end = pd.to_datetime(minute_end) + pd.Timedelta(hours=1)
    # 生成小时列表
    hour_list = pd.date_range(start=minute_start, end=new_minute_end, freq='h').strftime('%Y-%m-%d %H').tolist()
    # 移除多余的最后一项（因为多加了一小时）
    if hour_list and hour_list[-1] == (pd.to_datetime(minute_end) + pd.Timedelta(hours=1)).strftime('%Y-%m-%d %H'):
        hour_list = hour_list[:-1]

    return minute_list, hour_list

def count_objTweets_By_minhour(result_sheet, minhour='minute'):
    main_df = pd.read_excel(main_excel_file, sheet_name='MainData')
    # 筛选 mark_obj 列值为 1 的数据
    mobj_df = main_df[main_df['mark_obj'] == 1]

    tweetid_list = mobj_df['tweet_ID'].tolist()
    twttime_list = pd.to_datetime(mobj_df['tweet_pub_time']).dt.strftime('%Y-%m-%d %H:%M:%S')
    liulang_list = mobj_df['tweet_views'].tolist()
    pinglun_list = mobj_df['tweet_replies'].tolist()
    dianzan_list = mobj_df['tweet_likes'].tolist()
    zhuanfa_list = mobj_df['tweet_reposts'].tolist()
    stances_list = mobj_df['stance_new'].tolist()

    # 自动获取时间列表
    minute_list, hour_list = get_minute_hour_list(hic_minute_start, hic_minute_end)

    # 加载已有的工作簿
    wb = load_workbook(main_excel_file)
    # 检查是否存在名为 objnews_statis_minu 的工作表
    if result_sheet in wb.sheetnames:
        # 删除存在的工作表
        ws_to_delete = wb[result_sheet]
        wb.remove(ws_to_delete)
    # 创建新的工作表
    ws = wb.create_sheet(result_sheet)

    column_names = ['时间', "推文数_加权1", "浏览量_加权1", "评论量_加权1", "点赞量_加权1", "转发量_加权1",
                    "推文数_加权2", "浏览量_加权2", "评论量_加权2", "点赞量2_加权", "转发量_加权2"]
    ws.append(column_names)

    if minhour == 'minute':
        time_list = minute_list
    else:
        time_list = hour_list
    for timestr in time_list:
        print(timestr)
        result_list = list()
        result_list.append(timestr)
        newsheat = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  # 推文数_加权1 ...... 转发量_加权2

        for index, twettime in enumerate(twttime_list):
            if twettime.startswith(timestr):
                stance_str = str(stances_list[index]).lower()
                if ('hamas: very negative' in stance_str):
                    stance_value1 = -1
                    stance_value2 = -2
                elif ('hamas: mild negative' in stance_str):
                    stance_value1 = -1
                    stance_value2 = -1
                elif ('hamas: mild positive' in stance_str):
                    stance_value1 = 1
                    stance_value2 = 1
                elif ('hamas: very positive' in stance_str):
                    stance_value1 = 1
                    stance_value2 = 2
                else:
                    stance_value1 = 0
                    stance_value2 = 0

                newsheat[0] += stance_value1
                newsheat[1] += int(liulang_list[index]) * stance_value1
                newsheat[2] += int(pinglun_list[index]) * stance_value1
                newsheat[3] += int(dianzan_list[index]) * stance_value1
                newsheat[4] += int(zhuanfa_list[index]) * stance_value1
                newsheat[5] += stance_value2
                newsheat[6] += int(liulang_list[index]) * stance_value2
                newsheat[7] += int(pinglun_list[index]) * stance_value2
                newsheat[8] += int(dianzan_list[index]) * stance_value2
                newsheat[9] += int(zhuanfa_list[index]) * stance_value2

        for value in newsheat:
            result_list.append(value)

        ws.append(result_list)

    wb.save(main_excel_file)
    
def count_subjtweets_by_stance_minhour(subjtweets_sheet, stance_count_result_filename, minhour='minute'):
    data_frame = pd.read_excel(main_excel_file, sheet_name = subjtweets_sheet)
    twttime_list = data_frame['tweet_pub_time'].tolist()
    liulang_list = data_frame['tweet_views'].tolist()
    pinglun_list = data_frame['tweet_replies'].tolist()
    dianzan_list = data_frame['tweet_likes'].tolist()
    zhuanfa_list = data_frame['tweet_reposts'].tolist()
    stances_list = data_frame['stance_score'].tolist()

    minute_list, hour_list = get_minute_hour_list(hic_minute_start, hic_minute_end)

    wb = Workbook()
    ws = wb.active
    column_names = ['时间', "1_推文数", "1_浏览量", "1_评论量", "1_点赞量", "1_转发量", "2_推文数", "2_浏览量", "2_评论量",
                    "2_点赞量", "2_转发量", "3_推文数", "3_浏览量", "3_评论量", "3_点赞量", "3_转发量", "4_推文数", "4_浏览量",
                    "4_评论量", "4_点赞量", "4_转发量", "5_推文数", "5_浏览量", "5_评论量", "5_点赞量", "5_转发量"]
    ws.append(column_names)

    if minhour=='minute':
        minhour_list = minute_list
    else:
        minhour_list = hour_list
    for minhour_str in minhour_list:
        print(minhour_str)
        temp_list = list()
        temp_list.append(minhour_str)
        minute_stance_array = [[0, 0, 0, 0, 0], # very negative
                               [0, 0, 0, 0, 0], # mild negative
                               [0, 0, 0, 0, 0], # neutral
                               [0, 0, 0, 0, 0], # mild positive
                               [0, 0, 0, 0, 0]] # very positive
        for index, tweet_time in enumerate(twttime_list):
            if tweet_time.startswith(minhour_str):
                minute_stance_array[int(str(stances_list[index]))-1][0] += 1
                minute_stance_array[int(str(stances_list[index]))-1][1] += int(liulang_list[index])
                minute_stance_array[int(str(stances_list[index]))-1][2] += int(pinglun_list[index])
                minute_stance_array[int(str(stances_list[index]))-1][3] += int(dianzan_list[index])
                minute_stance_array[int(str(stances_list[index]))-1][4] += int(zhuanfa_list[index])

        for value in np.array(minute_stance_array).flatten():
            temp_list.append(str(value))
        ws.append(temp_list)

    wb.save(stance_count_result_filename)

def count_tweets_stance_by_minhour(sheetname, tweetstance_count_result_filepath, minhour='minute'):
    data_frame = pd.read_excel(main_excel_file, sheet_name = sheetname)
    tweetid_list = data_frame['tweet_ID'].tolist()
    twttime_list = data_frame['tweet_pub_time'].tolist()
    stances_list = data_frame['stance_score'].tolist()

    minute_list, hour_list = get_minute_hour_list(hic_minute_start, hic_minute_end)
    
    tweetstance_count_result_file = open(tweetstance_count_result_filepath, 'w', encoding='utf-8')
    minhour_stance_Dict = dict()
    if minhour == 'minute':
        time_list = minute_list
    else:
        time_list = hour_list
    for timestr in time_list:
        print(timestr)
        minhour_stance_Dict[timestr] = [0, 0, 0, 0, 0] # very negative, mild negative, neutral, mild positive, very positive
        for index, tweet_time in enumerate(twttime_list):
            if tweet_time.startswith(timestr):
                minhour_stance_Dict[timestr][int(str(stances_list[index]))-1] += 1

    json_data = json.dumps(minhour_stance_Dict)
    tweetstance_count_result_file.write(json_data)

def count_tweets_stance_by_user_minhour(sheetname, tweetstance_count_result_filepath, minhour='minute'):
    data_frame = pd.read_excel(main_excel_file, sheet_name=sheetname)
    twttime_list = data_frame['tweet_pub_time'].tolist()
    acuntID_list = data_frame['account_ID'].tolist()
    stances_list = data_frame['stance_score'].tolist()

    user_minhour_stance_Dict = dict()
    current_accountID = ''
    current_account_dict = dict()
    if minhour == 'minute':
        length = 16
    else:
        length = 13

    for index, accountID in enumerate(acuntID_list):
        if not str(accountID).strip() == current_accountID:
            if len(current_account_dict) >= 2:
                user_minhour_stance_Dict[current_accountID] = get_stance_average(current_account_dict)

            current_accountID = str(accountID).strip()
            current_account_dict = dict()
            current_account_dict[twttime_list[index][0:length]] = [int(stances_list[index]), 1]
        else:
            current_minhour = twttime_list[index][0:length]
            if current_minhour in current_account_dict.keys():
                stance_score_num = current_account_dict[current_minhour]   # total score and total num in one minhour
                stance_score_num[1] += 1
                stance_score_num[0] = stance_score_num[0] + int(stances_list[index])
                current_account_dict[current_minhour] = stance_score_num
            else:
                current_account_dict[current_minhour] = [int(stances_list[index]), 1]

    if len(current_account_dict) >= 2:
        user_minhour_stance_Dict[current_accountID] = get_stance_average(current_account_dict)

    stance_count_result_file = open(tweetstance_count_result_filepath, 'w', encoding='utf-8')
    json_data = json.dumps(user_minhour_stance_Dict)
    stance_count_result_file.write(json_data)

def get_stance_average(current_account_dict):
    current_account_dict1 = dict()
    for key, value in current_account_dict.items():
        avg_value = round(float(value[0]) / float(value[1]))
        current_account_dict1[key] = avg_value

    return current_account_dict1

def newsheat_by_user_2tweets_interval(stance_counts_filename, objnews_statis_sheet, newsheat_result_filename):
    newsheat_time_interval_dict = dict()
    stance_counts_dict = json.load(open(stance_counts_filename, 'r'))
    objnews_statis_data = pd.read_excel(main_excel_file, sheet_name=objnews_statis_sheet, dtype=str)
    objnews_time_list = objnews_statis_data['时间'].tolist()

    #因为运行时间较慢，因此分成多个进程运行
    #accountID的首字母集合为：{0~9, A~Z, _, a~z}，共63个
    #分开运行，然后合并json文件即可。
    start_chars = {'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T'}
    for accountID, minhour_dict in stance_counts_dict.items():
        if accountID[0] not in start_chars:
            continue
        print(accountID)
        
        for index in range(0, len(minhour_dict.keys())-1):
            minhour1 = list(minhour_dict.keys())[index]
            minhour2 = list(minhour_dict.keys())[index+1]
            index_minhour1 = 0
            index_minhour2 = 0
            for index1 in range(0, len(objnews_time_list)):
                if objnews_time_list[index1].startswith(minhour1):
                    index_minhour1 = index1
                    break
            for index2 in range(index_minhour1, len(objnews_time_list)):
                if objnews_time_list[index2].startswith(minhour2):
                    index_minhour2 = index2
                    break

            newsheat_count_array = [[0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0]]
            lists = [list(), list(), list(), list(), list()] #推文数, 浏览量, 评论量, 点赞量, 转发量
            for index3 in range(index_minhour1, index_minhour2+1):
                for index4 in range(0, 5):
                    # lists[index4].append(int(objnews_statis_data.iloc[index3,index4 + 1]))   # 用加权1数据
                    lists[index4].append(int(objnews_statis_data.iloc[index3, index4 + 6]))  # 用加权2数据

            for index5 in range(0, 5):
                newsheat_count_array[index5][0] = int(np.max(lists[index5])) #最大值
                newsheat_count_array[index5][1] = round(np.median(lists[index5])) #中位数
                newsheat_count_array[index5][2] = round(np.percentile(lists[index5], 75))#75分位数
                newsheat_count_array[index5][3] = round(np.mean(lists[index5]))#均值

            newsheat_time_interval_dict[accountID + '_' + minhour1 + '_' + minhour2] = newsheat_count_array

    result_file = open(newsheat_result_filename, 'w', encoding='utf-8')
    json_data = json.dumps(newsheat_time_interval_dict)
    result_file.write(json_data)

def count_tweets_by_stance_views_minhour(tweets_sheetname, stance_views_count_result_filepath, minhour='minute'):
    data_frame = pd.read_excel(main_excel_file, sheet_name=tweets_sheetname, dtype=str)
    twttime_list = data_frame['tweet_pub_time'].tolist()
    stances_list = data_frame['stance_score'].tolist()

    liulang_list = data_frame['tweet_views'].tolist()
    pinglun_list = data_frame['tweet_replies'].tolist()
    dianzan_list = data_frame['tweet_likes'].tolist()
    zhuanfa_list = data_frame['tweet_reposts'].tolist()

    minute_list, hour_list = get_minute_hour_list(hic_minute_start, hic_minute_end)
    stance_count_result_file = open(stance_views_count_result_filepath, 'w', encoding='utf-8')
    minhour_stance_views_Dict = dict()
    if minhour == 'minute':
        minhour_list = minute_list
    else:
        minhour_list = hour_list
    for minhour_str in minhour_list:
        print(minhour_str)
        minhour_stance_views_Dict[minhour_str] = [[0, 0, 0, 0, 0], # tweets count, from 1 to 5
                                                  [0, 0, 0, 0, 0], # liulan count, from 1 to 5
                                                  [0, 0, 0, 0, 0], # pinglun count, from 1 to 5
                                                  [0, 0, 0, 0, 0], # dianzan count, from 1 to 5
                                                  [0, 0, 0, 0, 0]] # zhuanfa count, from 1 to 5
        for index, tweet_time in enumerate(twttime_list):
            if tweet_time.startswith(minhour_str):
                minhour_stance_views_Dict[minhour_str][0][int(str(stances_list[index]))-1] += 1 # tweets count
                minhour_stance_views_Dict[minhour_str][1][int(str(stances_list[index]))-1] += int(liulang_list[index]) # liulan count
                minhour_stance_views_Dict[minhour_str][2][int(str(stances_list[index]))-1] += int(pinglun_list[index]) # pinglun count
                minhour_stance_views_Dict[minhour_str][3][int(str(stances_list[index]))-1] += int(dianzan_list[index]) # dianzan count
                minhour_stance_views_Dict[minhour_str][4][int(str(stances_list[index]))-1] += int(zhuanfa_list[index]) # zhuanfa count

    json_data = json.dumps(minhour_stance_views_Dict)
    stance_count_result_file.write(json_data)

def subjtweets_by_stance_views_interval(stance_counts_user_minhour_filename, subjtweets_stance_views_filename, subjtweets_stance_views_minhour_interval_filename):
    subjtweets_stance_views_minhour_interval_dict = dict()

    stance_counts_user_minhour_dict = json.load(open(stance_counts_user_minhour_filename, 'r'))
    subjtweets_stance_views_dict = json.load(open(subjtweets_stance_views_filename, 'r'))
    subjtweets_minhours_list = list(subjtweets_stance_views_dict.keys())

    for accountID, minhours_dict in stance_counts_user_minhour_dict.items():
        print(accountID)
        for index in range(0, len(minhours_dict.keys())-1):
            minhour1 = list(minhours_dict.keys())[index]
            minhour2 = list(minhours_dict.keys())[index+1]
            index_minhour1 = 0
            index_minhour2 = 0

            stance_views_minute_interval_array = \
                [[[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0]],  # 推文数 -> veryneg, mildneg, neutral, mildpos, verypos -> max, median, 75%, average
                 [[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0]],  # 浏览量 -> veryneg, mildneg, neutral, mildpos, verypos -> max, median, 75%, average
                 [[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0]],  # 评论量 -> veryneg, mildneg, neutral, mildpos, verypos -> max, median, 75%, average
                 [[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0]],  # 点赞量 -> veryneg, mildneg, neutral, mildpos, verypos -> max, median, 75%, average
                 [[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0]]]  # 转发量 -> veryneg, mildneg, neutral, mildpos, verypos -> max, median, 75%, average

            for index1 in range(0, len(subjtweets_minhours_list)):
                if subjtweets_minhours_list[index1].startswith(minhour1):
                    index_minhour1 = index1
                    break
            for index2 in range(index_minhour1, len(subjtweets_minhours_list)):
                if subjtweets_minhours_list[index2].startswith(minhour2):
                    index_minhour2 = index2
                    break

            stances_views_lists = [[list(), list(), list(), list(), list()],  # 推文数 -> veryneg, mildneg, neutral, mildpos, verypos
                                   [list(), list(), list(), list(), list()],  # 浏览量 -> veryneg, mildneg, neutral, mildpos, verypos
                                   [list(), list(), list(), list(), list()],  # 评论量 -> veryneg, mildneg, neutral, mildpos, verypos
                                   [list(), list(), list(), list(), list()],  # 点赞量 -> veryneg, mildneg, neutral, mildpos, verypos
                                   [list(), list(), list(), list(), list()]]  # 转发量 -> veryneg, mildneg, neutral, mildpos, verypos
            for index3 in range(index_minhour1, index_minhour2+1):
                stance_views_array = subjtweets_stance_views_dict[subjtweets_minhours_list[index3]]
                for index4 in range(0, 5):
                    for index5 in range(0, 5):
                        stances_views_lists[index4][index5].append(int(stance_views_array[index4][index5]))

            for index6 in range(0, 5):
                for index7 in range(0, 5):
                    stance_views_minute_interval_array[index6][index7][0] = int(np.max(stances_views_lists[index6][index7]))                 # 最大值
                    stance_views_minute_interval_array[index6][index7][1] = round(np.median(stances_views_lists[index6][index7]))            # 中位数
                    stance_views_minute_interval_array[index6][index7][2] = round(np.percentile(stances_views_lists[index6][index7], 75)) # 75分位数
                    stance_views_minute_interval_array[index6][index7][3] = round(np.mean(stances_views_lists[index6][index7]))              # 均值

            subjtweets_stance_views_minhour_interval_dict[accountID+'_'+minhour1+'_'+minhour2] = stance_views_minute_interval_array

    result_file = open(subjtweets_stance_views_minhour_interval_filename, 'w', encoding='utf-8')
    json_data = json.dumps(subjtweets_stance_views_minhour_interval_dict)
    result_file.write(json_data)

def softmax_sf(array1, sf=10.0): # sf = smooth factor
    total_sum = 0.0
    for index in range(0, len(array1)):
        total_sum += math.exp(float(0-array1[index])/sf)

    for index in range(0, len(array1)):
        array1[index] = round(math.exp(float(0-array1[index])/sf) / total_sum, 6)

    return array1

def stance_distance_user_interval_firsttime(stance_counts_user_minute_filename, stance_distance_user_interval_firsttime_filename, if_pow2=False):
    stance_distance_user_interval_firsttime_dict = dict()
    stance_counts_user_minute_dict = json.load(open(stance_counts_user_minute_filename, 'r'))

    for accountID, minutes_dict in stance_counts_user_minute_dict.items():
        print(accountID)
        for index in range(0, len(minutes_dict.keys())-1):
            minute1 = list(minutes_dict.keys())[index]
            minute2 = list(minutes_dict.keys())[index+1]
            stance1 = int(list(minutes_dict.values())[index])

            stance_distance_user_interval_firsttime_array = [0,0,0,0,0]  # distance of very neg, mild neg, neu, mild pos, very pos

            for index1 in range(0, 5):
                if ((stance1<=2 and index1>=3) or (stance1>=4 and index1<=1)) and if_pow2:
                    stance_distance_user_interval_firsttime_array[index1] = math.pow(index1+1-stance1, 2)
                else:
                    stance_distance_user_interval_firsttime_array[index1] = abs(index1+1-stance1)

            stance_distance_user_interval_firsttime_array = softmax_sf(stance_distance_user_interval_firsttime_array, 10)
            stance_distance_user_interval_firsttime_dict[accountID+'_'+minute1+'_'+minute2] = stance_distance_user_interval_firsttime_array

    result_file = open(stance_distance_user_interval_firsttime_filename, 'w', encoding='utf-8')
    json_data = json.dumps(stance_distance_user_interval_firsttime_dict)
    result_file.write(json_data)

#gij_type   = 0/1/2/3/4,    tweetsum, liulan, pinglun, dianzan, zhuanfa
#gij_value  = 0/1/2/3,      max, median, 75-percentile, average
#en_type    = 0/1/2/3/4,    tweetsum, liulan, pinglun, dianzan, zhuanfa
#en_value   = 0/1/2/3,      max, median, 75-percentile, average
def Linear_fit_stance_change(stance_counts_user_minute_filename, tweets_stance_views_minute_interval_filename,
                             stance_distance1_user_interval_filename, stance_distance2_user_interval_filename,
                             newsheat_user_minute_interval_filename, gij_type, gij_value, en_type, en_value):
    B_n_list = list()
    B_n1_list = list()
    E_n_list = list()
    g_ij = [list(),list(),list(),list(),list()] #veryneg, mildneg, neutral, mildpos, verypos
    r_ij = [list(),list(),list(),list(),list()] #veryneg, mildneg, neutral, mildpos, verypos
    a_ij = [list(),list(),list(),list(),list()] #veryneg, mildneg, neutral, mildpos, verypos

    stance_counts_user_minute_dict = json.load(open(stance_counts_user_minute_filename, 'r'))
    tweets_stance_views_minute_interval_dict = json.load(open(tweets_stance_views_minute_interval_filename, 'r'))
    stance_distance1_user_interval_dict = json.load(open(stance_distance1_user_interval_filename, 'r'))
    stance_distance2_user_interval_dict = json.load(open(stance_distance2_user_interval_filename, 'r'))
    newsheat_user_minute_interval_dict = json.load(open(newsheat_user_minute_interval_filename, 'r'))

    for accountID, minutes_dict in stance_counts_user_minute_dict.items():
        for index in range(0, len(minutes_dict.keys())-1):
            minute1 = list(minutes_dict.keys())[index]
            minute2 = list(minutes_dict.keys())[index + 1]
            stance1 = list(minutes_dict.values())[index]
            stance2 = list(minutes_dict.values())[index+1]
            B_n_list.append(float(stance1))
            B_n1_list.append(float(stance2))

            for index1 in range(0, 5):
                g_ij[index1].append(float(tweets_stance_views_minute_interval_dict[accountID+'_'+minute1+'_'+minute2][gij_type][index1][gij_value]))
                r_ij[index1].append(float(stance_distance1_user_interval_dict[accountID+'_'+minute1+'_'+minute2][index1]))
                a_ij[index1].append(float(stance_distance2_user_interval_dict[accountID+'_'+minute1+'_'+minute2][index1]))

            E_n_list.append(float(newsheat_user_minute_interval_dict[accountID+'_'+minute1+'_'+minute2][en_type][en_value]))

    B_n = np.array(B_n_list)    # B_i(n)
    B_n1 = np.array(B_n1_list)  # B_i(n+1)
    E_n = np.array(E_n_list)

    # Calculate Y and X2
    Y = B_n1 - B_n
    X1_list = list()
    for index2 in range(0, len(B_n)):
        sum = 0.0
        for index3 in range(0, 5):
            sum += g_ij[index3][index2]*r_ij[index3][index2]*a_ij[index3][index2]*(float(index3+1)-B_n[index2])
        X1_list.append(sum)

    X1 = np.array(X1_list)
    #X2 = np.array([np.sum(g_ij[i] * r_ij[i] * a_ij[i] * (np.arange(1, 6) - B_n[i])) for i in range(len(B_n))])
    X2 = E_n

    resultfile = open(pDir+'/HIC//model_fitting_results//fit_result_'+str(gij_type)+'_'+str(gij_value)+'_'+str(en_type)+'_'+str(en_value)+'.txt', 'w', encoding='utf-8')

    #################################
    ##########  curve_fit  ##########
    #################################
    # Define the model function
    def model(X, alpha, beta):
        X1, X2 = X
        return alpha * X1 + beta * X2

    # Prepare the data for curve fitting
    X_data = np.vstack((X1, X2))

    # Initial guess for the parameters
    initial_guess = [1.0, 1.0]

    # Perform the curve fit
    params, covariance = curve_fit(model, X_data, Y, p0=initial_guess)

    alpha, beta = params

    # Calculate the predicted values
    Y_pred = model(X_data, alpha, beta)

    # Calculate R^2
    residuals = Y - Y_pred
    ss_res = np.sum(residuals ** 2)
    ss_tot = np.sum((Y - np.mean(Y)) ** 2)
    r_squared = 1 - (ss_res / ss_tot)

    # Calculate RMSE
    rmse = np.sqrt(np.mean(residuals ** 2))

    # Calculate MAE
    mae = np.mean(np.abs(residuals))

    resultfile.write(f"*** curve_fit results ***\n")
    resultfile.write(f"Estimated alpha: {alpha}\n")
    resultfile.write(f"Estimated beta: {beta}\n")
    resultfile.write(f"R-squared: {r_squared}\n")
    resultfile.write(f"RMSE: {rmse}\n")
    resultfile.write(f"MAE: {mae}\n\n")

    #################################
    #########  statsmodels  #########
    #################################
    # Add a constant to the independent variables for intercept
    X = np.column_stack((X1, X2))
    X = sm.add_constant(X)

    # Fit the model using statsmodels
    ols_model = sm.OLS(Y, X).fit()

    # Print the summary to get p-values
    resultfile.write(f"*** statsmodel results ***\n")
    resultfile.write(str(ols_model.summary()))

def fit_result_analysis(fit_result_filedir, all_fit_results_filename):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    column_names = ['',"curve_fit_R-squared","curve_fit_RMSE","curve_fit_MAE","statsmodel_R-squared","statsmodel_Adj. R-squared",
                    "coef_const","std err_const","t_const","p_const","[0.025_const","0.975]_const","coef_X1","std err_X1",
                    "t_X1","p_X1","[0.025_X1","0.975]_X1","coef_X2","std err_X2","t_X2","p_X2","[0.025_X2","0.975]_X2"]
    ws.append(column_names)

    for filename in os.listdir(fit_result_filedir):
        result_list = list()
        result_list.append(filename)

        resultfilestr = open(fit_result_filedir + filename, 'r', encoding='utf-8').readlines()
        result_list.append(resultfilestr[3][resultfilestr[3].find('R-squared:')+len('R-squared:'):].strip())
        result_list.append(resultfilestr[4][resultfilestr[4].find('RMSE:')+len('RMSE:'):].strip())
        result_list.append(resultfilestr[5][resultfilestr[5].find('MAE:')+len('MAE:'):].strip() + '\n')

        result_list.append(resultfilestr[10][resultfilestr[10].find('R-squared:')+len('R-squared:'):].strip())
        result_list.append(resultfilestr[11][resultfilestr[11].find('Adj. R-squared:')+len('Adj. R-squared:'):].strip())

        for index in range(22, 25):
            str_line = resultfilestr[index]
            while '  ' in str_line:
                str_line = str_line.replace('  ', ' ')
            str_line_split = str_line.split(' ')
            for index in range(1, 7):
                result_list.append(str_line_split[index])

        ws.append(result_list)

    wb.save(all_fit_results_filename)

def newsheat_stance_by_user_interval(stance_counts_filename, 
                                     objnews_statis_sheetname,
                                     tweets_stance_views_minhour_interval_filename, 
                                     newsheat_result_filename, weight=1):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    column_names = ['',"stance1","stance2","tweetsnum_max","tweetsnum_median","tweetsnum_Q3","liulan_max",
                    "liulan_median","pinglun_max","pinglun_median","g_max_1","g_median_1","g_Q3_1","g_max_2",
                    "g_median_2","g_Q3_2","g_max_3","g_median_3","g_Q3_3","g_max_4","g_median_4","g_Q3_4",
                    "g_max_5","g_median_5","g_Q3_5"]
    ws.append(column_names)

    stance_counts_dict = json.load(open(stance_counts_filename, 'r'))
    objnews_statis_data = pd.read_excel(main_excel_file, sheet_name=objnews_statis_sheetname, dtype=str)
    objnews_time_list = objnews_statis_data['时间'].tolist()
    #open(objnews_statis_filename, 'r', encoding='utf-8').readlines()

    g_ij = [list(),list(),list(),list(),list()] #veryneg, mildneg, neutral, mildpos, verypos
    tweets_stance_views_minhour_interval_dict = json.load(open(tweets_stance_views_minhour_interval_filename, 'r'))

    for accountID, minhour_dict in stance_counts_dict.items():
        if accountID[0] not in accountID_init[0]:  # 运行速度慢，因此分为多个进程进行分析。
            continue
        print(accountID)
        for index in range(0, len(minhour_dict.keys())-1):
            result_list = list()

            minhour1 = list(minhour_dict.keys())[index]
            minhour2 = list(minhour_dict.keys())[index+1]
            result_list.append(accountID+'_'+minhour1+'_'+minhour2)

            stance1 = list(minhour_dict.values())[index]
            stance2 = list(minhour_dict.values())[index+1]
            result_list.append(stance1)
            result_list.append(stance2)

            index_minhour1 = 0
            index_minhour2 = 0
            for index1 in range(0, len(objnews_time_list)):
                if objnews_time_list[index1].startswith(minhour1[0:16]):
                    index_minhour1 = index1
                    break
            for index2 in range(index_minhour1, len(objnews_time_list)):
                if objnews_time_list[index2].startswith(minhour2[0:16]):
                    index_minhour2 = index2
                    break

            if index_minhour2 - index_minhour1 >= 0:
                lists1 = [list(), list(), list()]  # 推文数, 浏览量, 评论量
                for index3 in range(index_minhour1, index_minhour2+1):
                    line1_split = objnews_statis_data.iloc[index3]

                    for index4 in range(0, 3):
                        lists1[index4].append(int(line1_split[(weight-1)*5+index4+1]))

                result_list.append(int(np.max(lists1[0])))            #tweetsum的最大值
                result_list.append(int(np.median(lists1[0])))         #tweetsum的中位数
                result_list.append(int(np.percentile(lists1[0], 75))) #tweetsum的75分位
                for index5 in range(1, 3):
                    result_list.append(int(np.max(lists1[index5])))         #最大值
                    result_list.append(round(np.median(lists1[index5])))    #中位数


                gij_type = 0                       # tweetsum
                for stance_value in range(0, 5):    # veryneg, mildneg, neutral, mildpos, verypos
                    for gij_value in range(0, 3):   # gij_value  = 0/1/2,  max, median, 75-percentile, average
                        result_list.append(float(tweets_stance_views_minhour_interval_dict[accountID + '_' + minhour1 + '_' + minhour2][gij_type][stance_value][gij_value]))

            else:
                for index6 in range(0, 27):
                    result_list.append(0)

            ws.append(result_list)

    wb.save(newsheat_result_filename)

def filter_non_na_tweets():
    # 以字符串形式读取 Excel 文件
    df = pd.read_excel(main_excel_file, dtype=str)

    # 获取“stance_new”不包含 "hamas: na;" 的所有数据
    filtered_df = df[~df['stance_new'].str.contains("hamas: na;", na=False)]

    # 增加 stance_score 列
    def get_stance_score(stance_str):
        stance_str = stance_str.lower()
        if 'hamas: very positive' in stance_str:
            return 5
        elif 'hamas: mild positive' in stance_str:
            return 4
        elif 'hamas: neutral' in stance_str:
            return 3
        elif 'hamas: mild negative' in stance_str:
            return 2
        elif 'hamas: very negative' in stance_str:
            return 1
        else:
            return None

    filtered_df.loc[:, 'stance_score'] = filtered_df['stance_new'].apply(get_stance_score)

    # 检查无法确定得分的行
    for index, row in filtered_df[filtered_df['stance_score'].isnull()].iterrows():
        print(f"Index {index} has an unknown stance_new value: {row['stance_new']}")

    # 按照 'tweet_pub_time' 从早到晚进行排序
    sorted_by_time = filtered_df.sort_values(by='tweet_pub_time')

    # 按照 'account_ID' 字母表排序，再按照 'tweet_pub_time' 从早到晚排序
    sorted_by_user = filtered_df.sort_values(by=['account_ID', 'tweet_pub_time'])

    # 加载工作簿
    wb = load_workbook(main_excel_file)

    # 处理 "Tweets_wStance_SortByTime" 工作表
    if 'Tweets_wStance_SortByTime' in wb.sheetnames:
        del wb['Tweets_wStance_SortByTime']
    wb.create_sheet('Tweets_wStance_SortByTime')
    ws_time = wb['Tweets_wStance_SortByTime']
    # 写入表头
    ws_time.append(sorted_by_time.columns.tolist())
    # 写入数据
    for row in sorted_by_time.values.tolist():
        ws_time.append([str(cell) for cell in row])
    # 设置 "Tweets_wStance_SortByTime" 工作表单元格格式为文本
    for row in ws_time.iter_rows(min_row=2):
        for cell in row:
            cell.number_format = '@'

    # 处理 "Tweets_wStance_SortByUser" 工作表
    if 'Tweets_wStance_SortByUser' in wb.sheetnames:
        del wb['Tweets_wStance_SortByUser']
    wb.create_sheet('Tweets_wStance_SortByUser')
    ws_user = wb['Tweets_wStance_SortByUser']
    # 写入表头
    ws_user.append(sorted_by_user.columns.tolist())
    # 写入数据
    for row in sorted_by_user.values.tolist():
        ws_user.append([str(cell) for cell in row])
    # 设置 "Tweets_wStance_SortByUser" 工作表单元格格式为文本
    for row in ws_user.iter_rows(min_row=2):
        for cell in row:
            cell.number_format = '@'

    # 保存工作簿
    wb.save(main_excel_file)

def filter_subjboth_tweets():
    # 以字符串形式读取 Excel 文件
    df = pd.read_excel(main_excel_file, sheet_name='MainData', dtype=str)
    # 筛选 stance 列不包含 ;;na;; 但包含 subjective;; 或者 both;; 的数据
    filtered_df = df[
        (~df['stance_new'].str.contains('hamas: na;', na=False)) &
        ((df['stance_new'].str.contains('subjective', na=False)) | 
         (df['stance_new'].str.contains('both', na=False)))
    ]

    # 增加 stance_score 列
    def get_stance_score(stance_str):
        stance_str = stance_str.lower()
        if 'hamas: very positive' in stance_str:
            return 5
        elif 'hamas: mild positive' in stance_str:
            return 4
        elif 'hamas: neutral' in stance_str:
            return 3
        elif 'hamas: mild negative' in stance_str:
            return 2
        elif 'hamas: very negative' in stance_str:
            return 1
        else:
            return None

    filtered_df.loc[:, 'stance_score'] = filtered_df['stance_new'].apply(get_stance_score)

    # 检查无法确定得分的行
    for index, row in filtered_df[filtered_df['stance_score'].isnull()].iterrows():
        print(f"Index {index} has an unknown stance_new value: {row['stance_new']}")

    # # 按照 'tweet_pub_time' 从早到晚进行排序
    # sorted_by_time = filtered_df.sort_values(by='tweet_pub_time')

    # # 按照 'account_ID' 字母表排序，再按照 'tweet_pub_time' 从早到晚排序
    # sorted_by_user = filtered_df.sort_values(by=['account_ID', 'tweet_pub_time'])

    # 加载工作簿
    wb = load_workbook(main_excel_file)

    # 处理 "Tweets_subjboth" 工作表
    if 'Tweets_subjboth' in wb.sheetnames:
        del wb['Tweets_subjboth']
    wb.create_sheet('Tweets_subjboth')
    ws_time = wb['Tweets_subjboth']
    # 写入表头
    ws_time.append(filtered_df.columns.tolist())
    # 写入数据
    for row in filtered_df.values.tolist():
        ws_time.append([str(cell) for cell in row])
    # 设置 "Tweets_subjboth" 工作表单元格格式为文本
    for row in ws_time.iter_rows(min_row=2):
        for cell in row:
            cell.number_format = '@'

    # 保存工作簿
    wb.save(main_excel_file)

if __name__ == '__main__':

    ### [1] 新闻相关推文(objective-tweets)的统计：

    # step 1: 计算tweets与新闻（0.HAMAS_News）的相关度，并选出相关度高的objtweets。
    # 在excel文件里进行标记，列名为 'mark_obj'。
    # 在mark_objtweets_by_news-sim_HIC.py中实现（需要GPU服务器）。
    
    # step 2: 按分钟统计时间内的各项数据，
    # 存入 main_excel_file 的 名为'objnews_statis_minu'的sheet
    #count_objTweets_By_minhour('objnews_statis_minu', 'minute')

    ### [2] 非NA-Stance的tweets的分析：：

    # step 3: 筛选stance非na的tweets，
    # 结果分别存入sheet "Tweets_wStance_SortByTime"，
    # 以及 "Tweets_wStance_SortByUser"
    #filter_non_na_tweets()

    # step 4: 按分钟统计tweets的 stance 分布
    #count_tweets_stance_by_minhour('Tweets_wStance_SortByTime', pDir+'/HIC/TweetStance_Counts_byMinu.json', 'minute')

    # step 5: 按user/分钟统计tweets的 stance 分布
    #count_tweets_stance_by_user_minhour('Tweets_wStance_SortByUser', pDir+'/HIC/TweetStance_Counts_byUserMinu.json', 'minute')

    # step 6: 计算user两条推文之间的 news heat（新闻热度）
    #newsheat_by_user_2tweets_interval(pDir+'/HIC/TweetStance_Counts_byUserMinu.json', 'objnews_statis_minu', pDir+'/HIC/Newsheat_byUser2TweetsInterval_minu.json')

    ### [3] subjective-tweets 的统计：
    # step 7: 筛选 subj and both 的tweets，
    # 结果分别存入sheet "Tweets_subjboth"
    #filter_subjboth_tweets()

    # step 8. 按 分钟 对subj推文进行统计。计算每一小时内的
    # 五个类别的推文的推文数/浏览量/评论量/点赞数/转发量。
    #count_tweets_by_stance_views_minhour('Tweets_subjboth', pDir+'/HIC/SubjBothTweets_byStanceViewsMinu.json', 'minute')

    # step 9. 对用户两次 tweets 之间的 subj推文 进行分析。
    # 计算 五个类别的推文的推文数/浏览量/评论量/点赞数/转发量 的 
    # 最大值/中位数/75分位数/平均值。
    # subjtweets_by_stance_views_interval(pDir+'/HIC/TweetStance_Counts_byUserMinu.json',
    #                                     pDir+'/HIC/SubjBothTweets_byStanceViewsMinu.json',
    #                                     pDir+'/HIC/SubjBothTweets_byStanceViewsInterval_minu.json')

    # step 10. 对用户两次发言之前的新闻指数进行统计，
    # 计算两次发言之间的新闻/交叉新闻指数：推文数/浏览量/评论数的最大值/中位数。
    # newsheat_stance_by_user_interval(pDir + '/HIC/TweetStance_Counts_byUserMinu.json',
    #                                  'objnews_statis_minu',
    #                                  pDir + '/HIC/SubjBothTweets_byStanceViewsInterval_minu.json',
    #                                  pDir + '/HIC/Newsheat_Stance_byUserInterval_2_minu_0.xlsx', 2)

    # step 11. 对subjtweets进行分析，按分钟统计 每种stance 的 推文数，浏览量，评论数
    count_subjtweets_by_stance_minhour('Tweets_subjboth', pDir+'/HIC/subjtweets_statis_minu.xlsx', 'minute')